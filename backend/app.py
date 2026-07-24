import json
import os
import platform
import sqlite3

from collections import Counter
from datetime import datetime
from functools import wraps
from importlib.metadata import PackageNotFoundError, version
from io import BytesIO
from pathlib import Path
from uuid import uuid4

from flask import (
    Flask,
    flash,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for
)

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle
)

from ai_engine import (
    find_answer_details,
    get_questions_by_category,
    load_questions
)

from database import (
    add_chat,
    clear_all_chats,
    get_all_chats,
    init_database,
    update_feedback,
    update_rating
)


# ==================================================
# PROJE DOSYA YOLLARI
# ==================================================

PROJECT_FOLDER = Path(__file__).resolve().parent.parent

TEMPLATE_FOLDER = (
    PROJECT_FOLDER
    / "frontend"
    / "templates"
)

STATIC_FOLDER = (
    PROJECT_FOLDER
    / "frontend"
    / "static"
)

QUESTIONS_FILE = (
    PROJECT_FOLDER
    / "data"
    / "questions.json"
)

DATABASE_FILE = (
    PROJECT_FOLDER
    / "data"
    / "smart_support.db"
)

SETTINGS_FILE = (
    PROJECT_FOLDER
    / "data"
    / "settings.json"
)


# ==================================================
# KATEGORİLER
# ==================================================

CATEGORIES = {
    "internet": "🌐 İnternet",
    "ses": "🔊 Ses",
    "yazici": "🖨️ Yazıcı",
    "mikrofon": "🎤 Mikrofon",
    "donanim": "🖥️ Donanım",
    "usb": "🔌 USB",
    "kamera": "📷 Kamera",
    "diger": "🛠️ Diğer"
}


# ==================================================
# VARSAYILAN YÖNETİCİ BİLGİLERİ
# ==================================================

DEFAULT_ADMIN_USERNAME = os.getenv(
    "ADMIN_USERNAME",
    "admin"
)

DEFAULT_ADMIN_PASSWORD = os.getenv(
    "ADMIN_PASSWORD",
    "smart123"
)


# ==================================================
# FLASK UYGULAMASI
# ==================================================

app = Flask(
    __name__,
    template_folder=str(TEMPLATE_FOLDER),
    static_folder=str(STATIC_FOLDER)
)

app.secret_key = os.getenv(
    "SECRET_KEY",
    "smart-support-ai-gizli-anahtar"
)


# Veritabanı tablolarını oluşturur.
init_database()


# ==================================================
# YARDIMCI FONKSİYONLAR
# ==================================================

def admin_login_required(view_function):
    """
    Yönetici girişi yapılmadan admin
    sayfalarına erişilmesini engeller.
    """

    @wraps(view_function)
    def wrapped_view(*args, **kwargs):

        if not session.get("admin_logged_in"):

            flash(
                "Yönetici paneline erişmek için giriş yapmalısınız.",
                "error"
            )

            return redirect(
                url_for("admin_login")
            )

        return view_function(
            *args,
            **kwargs
        )

    return wrapped_view


def save_questions(support_items):
    """
    Bilgi tabanını questions.json
    dosyasına kaydeder.
    """

    try:

        QUESTIONS_FILE.parent.mkdir(
            parents=True,
            exist_ok=True
        )

        with QUESTIONS_FILE.open(
            "w",
            encoding="utf-8"
        ) as file:

            json.dump(
                support_items,
                file,
                ensure_ascii=False,
                indent=4
            )

        return True

    except OSError as error:

        print(
            "Bilgi tabanı kaydedilemedi:",
            error
        )

        return False


def save_settings(settings):
    """
    Yönetici ayarlarını settings.json
    dosyasına kaydeder.
    """

    try:

        SETTINGS_FILE.parent.mkdir(
            parents=True,
            exist_ok=True
        )

        with SETTINGS_FILE.open(
            "w",
            encoding="utf-8"
        ) as file:

            json.dump(
                settings,
                file,
                ensure_ascii=False,
                indent=4
            )

        return True

    except OSError as error:

        print(
            "Ayarlar kaydedilemedi:",
            error
        )

        return False


def load_settings():
    """
    Yönetici kullanıcı adı ve şifresini
    settings.json dosyasından okur.
    """

    default_settings = {
        "admin_username": DEFAULT_ADMIN_USERNAME,
        "admin_password": DEFAULT_ADMIN_PASSWORD
    }

    if not SETTINGS_FILE.exists():

        save_settings(
            default_settings
        )

        return default_settings

    try:

        with SETTINGS_FILE.open(
            "r",
            encoding="utf-8"
        ) as file:

            settings = json.load(
                file
            )

        if not isinstance(settings, dict):

            save_settings(
                default_settings
            )

            return default_settings

        settings.setdefault(
            "admin_username",
            DEFAULT_ADMIN_USERNAME
        )

        settings.setdefault(
            "admin_password",
            DEFAULT_ADMIN_PASSWORD
        )

        return settings

    except (
        OSError,
        json.JSONDecodeError
    ) as error:

        print(
            "Ayarlar okunamadı:",
            error
        )

        return default_settings


def get_system_versions():
    """
    Python, Flask ve SQLite sürümlerini
    döndürür.
    """

    try:

        flask_version = version(
            "Flask"
        )

    except PackageNotFoundError:

        flask_version = (
            "Tespit edilemedi"
        )

    return {
        "python": platform.python_version(),
        "flask": flask_version,
        "sqlite": sqlite3.sqlite_version,
        "operating_system": platform.system(),
        "platform": platform.platform()
    }


def validate_imported_questions(imported_data):
    """
    Yüklenen JSON bilgi tabanını kontrol eder.
    """

    if not isinstance(
        imported_data,
        list
    ):

        return (
            False,
            "JSON dosyasının ana yapısı liste olmalıdır.",
            []
        )

    validated_items = []

    for item_number, item in enumerate(
        imported_data,
        start=1
    ):

        if not isinstance(
            item,
            dict
        ):

            return (
                False,
                "{}. kayıt geçerli değildir.".format(
                    item_number
                ),
                []
            )

        category = str(
            item.get(
                "category",
                ""
            )
        ).strip()

        question = str(
            item.get(
                "question",
                ""
            )
        ).strip()

        answer = str(
            item.get(
                "answer",
                ""
            )
        ).strip()

        keywords = item.get(
            "keywords",
            []
        )

        if category not in CATEGORIES:

            return (
                False,
                "{}. kayıtta geçersiz kategori bulunuyor.".format(
                    item_number
                ),
                []
            )

        if not question:

            return (
                False,
                "{}. kayıtta soru alanı boş.".format(
                    item_number
                ),
                []
            )

        if not answer:

            return (
                False,
                "{}. kayıtta cevap alanı boş.".format(
                    item_number
                ),
                []
            )

        if isinstance(
            keywords,
            str
        ):

            keywords = [
                keyword.strip()
                for keyword in keywords.split(",")
                if keyword.strip()
            ]

        if not isinstance(
            keywords,
            list
        ):

            keywords = []

        cleaned_keywords = [
            str(keyword).strip()
            for keyword in keywords
            if str(keyword).strip()
        ]

        if not cleaned_keywords:

            cleaned_keywords = [
                question
            ]

        validated_items.append(
            {
                "category": category,
                "question": question,
                "keywords": cleaned_keywords,
                "answer": answer
            }
        )

    return (
        True,
        "Bilgi tabanı geçerli.",
        validated_items
    )


def convert_chat_to_dictionary(chat):
    """
    Veritabanından gelen sohbet kaydını
    sözlük biçimine dönüştürür.
    """

    if isinstance(chat, dict):
        return chat

    try:
        return dict(chat)

    except (
        TypeError,
        ValueError
    ):
        return {}


def prepare_chat_history(chat_history):
    """
    Sohbet kayıtlarını sözlük listesine
    dönüştürür.
    """

    return [
        convert_chat_to_dictionary(chat)
        for chat in chat_history
    ]


def get_chat_value(
    chat,
    possible_keys,
    default="-"
):
    """
    Olası sütun isimlerinden bulunan
    ilk değeri döndürür.
    """

    chat_dictionary = (
        convert_chat_to_dictionary(chat)
    )

    for key in possible_keys:

        value = chat_dictionary.get(
            key
        )

        if (
            value is not None
            and value != ""
        ):

            return value

    return default


def calculate_statistics(chat_history):
    """
    Sohbet geçmişinden genel istatistikler ve
    yapay zekâ performans bilgileri üretir.
    """

    prepared_history = prepare_chat_history(
        chat_history
    )

    total_questions = len(
        prepared_history
    )

    positive_feedback = sum(
        1
        for chat in prepared_history
        if chat.get("feedback") == "positive"
    )

    negative_feedback = sum(
        1
        for chat in prepared_history
        if chat.get("feedback") == "negative"
    )

    unanswered_feedback = sum(
        1
        for chat in prepared_history
        if chat.get("feedback") not in [
            "positive",
            "negative"
        ]
    )

    category_counter = Counter(
        chat.get(
            "category",
            "diger"
        )
        for chat in prepared_history
    )

    if category_counter:

        most_used_category_key = (
            category_counter.most_common(1)[0][0]
        )

        most_used_category_count = (
            category_counter.most_common(1)[0][1]
        )

        most_used_category_name = CATEGORIES.get(
            most_used_category_key,
            CATEGORIES["diger"]
        )

    else:

        most_used_category_name = (
            "Henüz veri yok"
        )

        most_used_category_count = 0

    category_statistics = []

    for (
        category_key,
        category_name
    ) in CATEGORIES.items():

        category_statistics.append(
            {
                "key": category_key,
                "name": category_name,
                "count": category_counter.get(
                    category_key,
                    0
                )
            }
        )

    confidence_scores = []

    for chat in prepared_history:

        try:

            confidence_score = int(
                chat.get(
                    "confidence_score",
                    0
                )
            )

        except (
            TypeError,
            ValueError
        ):

            confidence_score = 0

        confidence_scores.append(
            confidence_score
        )

    if confidence_scores:

        average_confidence = round(
            sum(confidence_scores)
            / len(confidence_scores)
        )

    else:

        average_confidence = 0

    high_confidence_count = sum(
        1
        for score in confidence_scores
        if score >= 70
    )

    medium_confidence_count = sum(
        1
        for score in confidence_scores
        if 40 <= score < 70
    )

    low_confidence_count = sum(
        1
        for score in confidence_scores
        if score < 40
    )

    suggestion_count = sum(
        1
        for chat in prepared_history
        if chat.get("match_type") == "suggestion"
    )

    failed_match_count = sum(
        1
        for chat in prepared_history
        if chat.get("match_type") in [
            "not_found",
            "no_category_data",
            "empty"
        ]
    )

    low_confidence_questions = []

    for chat in reversed(
        prepared_history
    ):

        try:

            confidence_score = int(
                chat.get(
                    "confidence_score",
                    0
                )
            )

        except (
            TypeError,
            ValueError
        ):

            confidence_score = 0

        if confidence_score < 40:

            low_confidence_questions.append(
                {
                    "question": chat.get(
                        "question",
                        "-"
                    ),
                    "category_name": chat.get(
                        "category_name",
                        "Diğer"
                    ),
                    "confidence_score": confidence_score,
                    "match_type": chat.get(
                        "match_type",
                        "not_found"
                    ),
                    "feedback": chat.get(
                        "feedback"
                    )
                }
            )

    return {
        "total_questions": total_questions,
        "positive_feedback": positive_feedback,
        "negative_feedback": negative_feedback,
        "unanswered_feedback": unanswered_feedback,
        "most_used_category_name": (
            most_used_category_name
        ),
        "most_used_category_count": (
            most_used_category_count
        ),
        "category_statistics": (
            category_statistics
        ),
        "average_confidence": average_confidence,
        "high_confidence_count": high_confidence_count,
        "medium_confidence_count": medium_confidence_count,
        "low_confidence_count": low_confidence_count,
        "suggestion_count": suggestion_count,
        "failed_match_count": failed_match_count,
        "low_confidence_questions": (
            low_confidence_questions[:10]
        )
    }
def normalize_question_text(text):
    """
    Soruları karşılaştırmak için metni
    küçük harfe çevirip boşlukları düzenler.
    """

    return " ".join(
        str(text).strip().lower().split()
    )


def get_learning_center_items():
    """
    Olumsuz geri bildirim verilen ve henüz
    bilgi tabanına eklenmemiş soruları döndürür.
    """

    chat_history = prepare_chat_history(
        get_all_chats()
    )

    support_items = load_questions()

    existing_questions = {
        normalize_question_text(
            item.get(
                "question",
                ""
            )
        )
        for item in support_items
    }

    learning_items = []

    seen_questions = set()

    for chat in reversed(
        chat_history
    ):

        if chat.get("feedback") != "negative":
            continue

        question = str(
            chat.get(
                "question",
                ""
            )
        ).strip()

        if not question:
            continue

        normalized_question = (
            normalize_question_text(
                question
            )
        )

        if normalized_question in existing_questions:
            continue

        if normalized_question in seen_questions:
            continue

        seen_questions.add(
            normalized_question
        )

        category = str(
            chat.get(
                "category",
                "diger"
            )
        ).strip()

        if category not in CATEGORIES:
            category = "diger"

        learning_items.append(
            {
                "chat_id": str(
                    chat.get(
                        "chat_id",
                        chat.get(
                            "id",
                            ""
                        )
                    )
                ),
                "question": question,
                "answer": str(
                    chat.get(
                        "answer",
                        "-"
                    )
                ),
                "category": category,
                "category_name": CATEGORIES.get(
                    category,
                    CATEGORIES["diger"]
                ),
                "confidence_score": chat.get(
                    "confidence_score",
                    0
                ),
                "matched_question": chat.get(
                    "matched_question",
                    "-"
                ),
                "created_at": format_report_date(
                    chat.get(
                        "created_at",
                        "-"
                    )
                )
            }
        )

    return learning_items


def format_report_date(date_value):
    """
    Tarih bilgisini rapor formatına çevirir.
    """

    if not date_value:
        return "-"

    if isinstance(
        date_value,
        datetime
    ):

        return date_value.strftime(
            "%d.%m.%Y %H:%M"
        )

    date_text = str(
        date_value
    )

    date_formats = [
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%dT%H:%M:%S",
        "%d.%m.%Y %H:%M"
    ]

    for date_format in date_formats:

        try:

            parsed_date = datetime.strptime(
                date_text,
                date_format
            )

            return parsed_date.strftime(
                "%d.%m.%Y %H:%M"
            )

        except ValueError:
            continue

    try:

        parsed_date = datetime.fromisoformat(
            date_text.replace(
                "Z",
                "+00:00"
            )
        )

        return parsed_date.strftime(
            "%d.%m.%Y %H:%M"
        )

    except ValueError:

        return date_text


def format_feedback(feedback):
    """
    Geri bildirim değerini Türkçeye çevirir.
    """

    if feedback == "positive":
        return "Olumlu"

    if feedback == "negative":
        return "Olumsuz"

    return "Değerlendirilmedi"


def register_pdf_fonts():
    """
    PDF içerisinde Türkçe karakter
    desteği için yazı tipi kaydeder.
    """

    regular_font_paths = [
        r"C:\Windows\Fonts\arial.ttf",
        r"C:\Windows\Fonts\calibri.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"
    ]

    bold_font_paths = [
        r"C:\Windows\Fonts\arialbd.ttf",
        r"C:\Windows\Fonts\calibrib.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"
    ]

    regular_font_path = None
    bold_font_path = None

    for font_path in regular_font_paths:

        if os.path.exists(
            font_path
        ):

            regular_font_path = font_path
            break

    for font_path in bold_font_paths:

        if os.path.exists(
            font_path
        ):

            bold_font_path = font_path
            break

    regular_font_name = "Helvetica"
    bold_font_name = "Helvetica-Bold"

    try:

        if regular_font_path:

            pdfmetrics.registerFont(
                TTFont(
                    "SmartSupportRegular",
                    regular_font_path
                )
            )

            regular_font_name = (
                "SmartSupportRegular"
            )

        if bold_font_path:

            pdfmetrics.registerFont(
                TTFont(
                    "SmartSupportBold",
                    bold_font_path
                )
            )

            bold_font_name = (
                "SmartSupportBold"
            )

    except Exception as error:

        print(
            "PDF yazı tipi kaydedilemedi:",
            error
        )

    return (
        regular_font_name,
        bold_font_name
    )


# ==================================================
# ANA KULLANICI SAYFASI
# ==================================================

@app.route(
    "/",
    methods=[
        "GET",
        "POST"
    ]
)
def home():
    """
    Kullanıcının teknik destek sorusu
    sorduğu ana sayfa.
    """

    if request.method == "POST":

        question = request.form.get(
            "question",
            ""
        ).strip()

        category = request.form.get(
            "category",
            "diger"
        ).strip()

        if category not in CATEGORIES:
            category = "diger"

        if not question:

            flash(
                "Lütfen bir teknik destek sorusu yazın.",
                "error"
            )

            return redirect(
                url_for("home")
            )

        answer_details = find_answer_details(
            question,
            category
        )

        answer = str(
            answer_details.get(
                "answer",
                ""
            )
        )

        confidence_score = int(
            answer_details.get(
                "confidence_score",
                0
            )
        )

        confidence_level = str(
            answer_details.get(
                "confidence_level",
                "Yetersiz"
            )
        )

        matched_question = answer_details.get(
            "matched_question"
        )

        suggestion = answer_details.get(
            "suggestion"
        )

        match_type = str(
            answer_details.get(
                "match_type",
                "not_found"
            )
        )

        chat_id = str(
            uuid4()
        )

        add_chat(
            chat_id=chat_id,
            question=question,
            answer=answer,
            category=category,
            category_name=CATEGORIES[category],
            confidence_score=confidence_score,
            confidence_level=confidence_level,
            matched_question=matched_question,
            suggestion=suggestion,
            match_type=match_type
        )

        return redirect(
            url_for("home")
            + "#chat-"
            + chat_id
        )

    chat_history = prepare_chat_history(
        get_all_chats()
    )

    category_questions = {
        category_key: get_questions_by_category(
            category_key
        )
        for category_key in CATEGORIES
    }

    statistics = calculate_statistics(
        chat_history
    )

    return render_template(
        "index.html",
        chat_history=chat_history,
        categories=CATEGORIES,
        category_questions=category_questions,
        statistics=statistics
    )

# ==================================================
# YÖNETİCİ GİRİŞ VE ÇIKIŞ
# ==================================================

@app.route(
    "/admin/login",
    methods=[
        "GET",
        "POST"
    ]
)
def admin_login():
    """
    Yönetici giriş sayfası.
    """

    if session.get(
        "admin_logged_in"
    ):

        return redirect(
            url_for("admin_panel")
        )

    if request.method == "POST":

        username = request.form.get(
            "username",
            ""
        ).strip()

        password = request.form.get(
            "password",
            ""
        )

        settings = load_settings()

        if (
            username
            == settings["admin_username"]
            and password
            == settings["admin_password"]
        ):

            session.clear()

            session[
                "admin_logged_in"
            ] = True

            session[
                "admin_username"
            ] = username

            flash(
                "Yönetici paneline başarıyla giriş yaptınız.",
                "success"
            )

            return redirect(
                url_for("admin_panel")
            )

        flash(
            "Kullanıcı adı veya şifre yanlış.",
            "error"
        )

    return render_template(
        "admin_login.html"
    )


@app.route("/admin/logout")
def admin_logout():
    """
    Yönetici hesabından çıkış yapar.
    """

    session.clear()

    flash(
        "Yönetici hesabından çıkış yapıldı.",
        "success"
    )

    return redirect(
        url_for("admin_login")
    )


# ==================================================
# ADMIN DASHBOARD
# ==================================================

@app.route("/admin")
@admin_login_required
def admin_panel():
    """
    Yönetici dashboard sayfası.
    """

    support_items = load_questions()

    chat_history = prepare_chat_history(
        get_all_chats()
    )

    statistics = calculate_statistics(
        chat_history
    )

    settings = load_settings()

    return render_template(
        "admin/dashboard.html",
        active_page="dashboard",
        categories=CATEGORIES,
        total_items=len(support_items),
        statistics=statistics,
        admin_username=settings.get(
            "admin_username",
            "admin"
        )
    )


# ==================================================
# BİLGİ TABANI
# ==================================================

@app.route("/admin/knowledge")
@admin_login_required
def admin_knowledge():
    """
    Bilgi tabanı kayıtlarını gösterir.
    """

    support_items = load_questions()

    settings = load_settings()

    return render_template(
        "admin/knowledge.html",
        active_page="knowledge",
        support_items=support_items,
        categories=CATEGORIES,
        total_items=len(support_items),
        admin_username=settings.get(
            "admin_username",
            "admin"
        )
    )


@app.route(
    "/admin/add",
    methods=["POST"]
)
@admin_login_required
def add_support_item():
    """
    Bilgi tabanına yeni kayıt ekler.
    """

    category = request.form.get(
        "category",
        ""
    ).strip()

    question = request.form.get(
        "question",
        ""
    ).strip()

    keywords_text = request.form.get(
        "keywords",
        ""
    ).strip()

    answer = request.form.get(
        "answer",
        ""
    ).strip()

    if category not in CATEGORIES:

        flash(
            "Lütfen geçerli bir kategori seçin.",
            "error"
        )

        return redirect(
            url_for("admin_knowledge")
        )

    if not question:

        flash(
            "Soru alanı boş bırakılamaz.",
            "error"
        )

        return redirect(
            url_for("admin_knowledge")
        )

    if not answer:

        flash(
            "Cevap alanı boş bırakılamaz.",
            "error"
        )

        return redirect(
            url_for("admin_knowledge")
        )

    keywords = [
        keyword.strip()
        for keyword in keywords_text.split(",")
        if keyword.strip()
    ]

    if not keywords:

        keywords = [
            question
        ]

    support_items = load_questions()

    question_lower = question.lower()

    duplicate_question = any(
        str(
            item.get(
                "question",
                ""
            )
        ).strip().lower() == question_lower
        for item in support_items
    )

    if duplicate_question:

        flash(
            "Bu soru bilgi tabanında zaten bulunuyor.",
            "error"
        )

        return redirect(
            url_for("admin_knowledge")
        )

    support_items.append(
        {
            "category": category,
            "question": question,
            "keywords": keywords,
            "answer": answer
        }
    )

    if not save_questions(
        support_items
    ):

        flash(
            "Kayıt eklenirken bir hata oluştu.",
            "error"
        )

        return redirect(
            url_for("admin_knowledge")
        )

    flash(
        "Yeni teknik destek kaydı başarıyla eklendi.",
        "success"
    )

    return redirect(
        url_for("admin_knowledge")
    )


@app.route(
    "/admin/edit/<int:item_index>",
    methods=["POST"]
)
@admin_login_required
def edit_support_item(item_index):
    """
    Bilgi tabanındaki kaydı günceller.
    """

    support_items = load_questions()

    if (
        item_index < 0
        or item_index >= len(
            support_items
        )
    ):

        flash(
            "Düzenlemek istediğiniz kayıt bulunamadı.",
            "error"
        )

        return redirect(
            url_for("admin_knowledge")
        )

    category = request.form.get(
        "category",
        ""
    ).strip()

    question = request.form.get(
        "question",
        ""
    ).strip()

    keywords_text = request.form.get(
        "keywords",
        ""
    ).strip()

    answer = request.form.get(
        "answer",
        ""
    ).strip()

    if category not in CATEGORIES:

        flash(
            "Lütfen geçerli bir kategori seçin.",
            "error"
        )

        return redirect(
            url_for("admin_knowledge")
        )

    if not question:

        flash(
            "Soru alanı boş bırakılamaz.",
            "error"
        )

        return redirect(
            url_for("admin_knowledge")
        )

    if not answer:

        flash(
            "Cevap alanı boş bırakılamaz.",
            "error"
        )

        return redirect(
            url_for("admin_knowledge")
        )

    keywords = [
        keyword.strip()
        for keyword in keywords_text.split(",")
        if keyword.strip()
    ]

    if not keywords:

        keywords = [
            question
        ]

    question_lower = question.lower()

    duplicate_question = any(
        index != item_index
        and str(
            item.get(
                "question",
                ""
            )
        ).strip().lower() == question_lower
        for index, item in enumerate(
            support_items
        )
    )

    if duplicate_question:

        flash(
            "Bu soru başka bir kayıtta zaten bulunuyor.",
            "error"
        )

        return redirect(
            url_for("admin_knowledge")
        )

    support_items[item_index] = {
        "category": category,
        "question": question,
        "keywords": keywords,
        "answer": answer
    }

    if not save_questions(
        support_items
    ):

        flash(
            "Kayıt düzenlenirken bir hata oluştu.",
            "error"
        )

        return redirect(
            url_for("admin_knowledge")
        )

    flash(
        "Teknik destek kaydı başarıyla güncellendi.",
        "success"
    )

    return redirect(
        url_for("admin_knowledge")
    )


@app.route(
    "/admin/delete/<int:item_index>",
    methods=["POST"]
)
@admin_login_required
def delete_support_item(item_index):
    """
    Bilgi tabanındaki kaydı siler.
    """

    support_items = load_questions()

    if (
        item_index < 0
        or item_index >= len(
            support_items
        )
    ):

        flash(
            "Silmek istediğiniz kayıt bulunamadı.",
            "error"
        )

        return redirect(
            url_for("admin_knowledge")
        )

    deleted_item = support_items.pop(
        item_index
    )

    if not save_questions(
        support_items
    ):

        flash(
            "Kayıt silinirken bir hata oluştu.",
            "error"
        )

        return redirect(
            url_for("admin_knowledge")
        )

    flash(
        "'{}' başarıyla silindi.".format(
            deleted_item.get(
                "question",
                "Kayıt"
            )
        ),
        "success"
    )

    return redirect(
        url_for("admin_knowledge")
    )
# ==================================================
# YAPAY ZEKÂ ÖĞRENME MERKEZİ
# ==================================================

@app.route("/admin/learning")
@admin_login_required
def admin_learning():
    """
    Olumsuz geri bildirim verilen ve bilgi
    tabanına eklenmesi gereken soruları gösterir.
    """

    learning_items = (
        get_learning_center_items()
    )

    settings = load_settings()

    return render_template(
        "admin/learning.html",
        active_page="learning",
        learning_items=learning_items,
        categories=CATEGORIES,
        total_learning_items=len(
            learning_items
        ),
        admin_username=settings.get(
            "admin_username",
            "admin"
        )
    )


@app.route(
    "/admin/learning/add",
    methods=["POST"]
)
@admin_login_required
def add_learning_item():
    """
    Öğrenme merkezindeki bir soruyu
    bilgi tabanına ekler.
    """

    category = request.form.get(
        "category",
        "diger"
    ).strip()

    question = request.form.get(
        "question",
        ""
    ).strip()

    keywords_text = request.form.get(
        "keywords",
        ""
    ).strip()

    answer = request.form.get(
        "answer",
        ""
    ).strip()

    if category not in CATEGORIES:
        category = "diger"

    if not question:

        flash(
            "Soru alanı boş bırakılamaz.",
            "error"
        )

        return redirect(
            url_for("admin_learning")
        )

    if not answer:

        flash(
            "Doğru cevap alanı boş bırakılamaz.",
            "error"
        )

        return redirect(
            url_for("admin_learning")
        )

    keywords = [
        keyword.strip()
        for keyword in keywords_text.split(",")
        if keyword.strip()
    ]

    if not keywords:

        keywords = [
            word
            for word in question.split()
            if len(word) >= 3
        ]

    if not keywords:

        keywords = [
            question
        ]

    support_items = load_questions()

    normalized_question = (
        normalize_question_text(
            question
        )
    )

    duplicate_question = any(
        normalize_question_text(
            item.get(
                "question",
                ""
            )
        ) == normalized_question
        for item in support_items
    )

    if duplicate_question:

        flash(
            "Bu soru bilgi tabanında zaten bulunuyor.",
            "error"
        )

        return redirect(
            url_for("admin_learning")
        )

    support_items.append(
        {
            "category": category,
            "question": question,
            "keywords": keywords,
            "answer": answer
        }
    )

    if not save_questions(
        support_items
    ):

        flash(
            "Öğrenilen cevap kaydedilemedi.",
            "error"
        )

        return redirect(
            url_for("admin_learning")
        )

    flash(
        "Yeni cevap bilgi tabanına eklendi. "
        "Sistem artık bu soruyu cevaplayabilir.",
        "success"
    )

    return redirect(
        url_for("admin_learning")
    )

# ==================================================
# RAPORLAR SAYFASI
# ==================================================

@app.route("/admin/reports")
@admin_login_required
def admin_reports():
    """
    Raporlar sayfasını gösterir.
    """

    chat_history = prepare_chat_history(
        get_all_chats()
    )

    statistics = calculate_statistics(
        chat_history
    )

    settings = load_settings()

    return render_template(
        "admin/reports.html",
        active_page="reports",
        categories=CATEGORIES,
        statistics=statistics,
        admin_username=settings.get(
            "admin_username",
            "admin"
        )
    )


# ==================================================
# EXCEL RAPORU
# ==================================================

@app.route("/admin/reports/excel")
@admin_login_required
def download_excel_report():
    """
    Sohbet geçmişini Excel raporu
    olarak indirir.
    """

    chat_history = prepare_chat_history(
        get_all_chats()
    )

    statistics = calculate_statistics(
        chat_history
    )

    workbook = Workbook()

    chat_sheet = workbook.active

    chat_sheet.title = (
        "Sohbet Geçmişi"
    )

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="4F46E5"
    )

    header_font = Font(
        color="FFFFFF",
        bold=True
    )

    title_font = Font(
        bold=True,
        size=16,
        color="312E81"
    )

    chat_sheet.merge_cells(
        "A1:G1"
    )

    chat_sheet["A1"] = (
        "Smart Support AI - Sohbet Raporu"
    )

    chat_sheet["A1"].font = (
        title_font
    )

    chat_sheet["A1"].alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    chat_sheet.row_dimensions[
        1
    ].height = 30

    chat_sheet["A2"] = (
        "Rapor Tarihi"
    )

    chat_sheet["B2"] = (
        datetime.now().strftime(
            "%d.%m.%Y %H:%M"
        )
    )

    headers = [
        "No",
        "Soru",
        "Cevap",
        "Kategori",
        "Geri Bildirim",
        "Tarih",
        "Kayıt Kimliği"
    ]

    header_row = 4

    for (
        column_index,
        header
    ) in enumerate(
        headers,
        start=1
    ):

        cell = chat_sheet.cell(
            row=header_row,
            column=column_index,
            value=header
        )

        cell.fill = (
            header_fill
        )

        cell.font = (
            header_font
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    for (
        row_index,
        chat
    ) in enumerate(
        chat_history,
        start=1
    ):

        excel_row = (
            header_row
            + row_index
        )

        feedback = format_feedback(
            get_chat_value(
                chat,
                ["feedback"],
                ""
            )
        )

        category = get_chat_value(
            chat,
            [
                "category_name",
                "category"
            ],
            "Diğer"
        )

        date_value = get_chat_value(
            chat,
            [
                "created_at",
                "date",
                "timestamp"
            ],
            "-"
        )

        chat_id = get_chat_value(
            chat,
            [
                "chat_id",
                "id"
            ],
            "-"
        )

        row_values = [
            row_index,
            get_chat_value(
                chat,
                ["question"]
            ),
            get_chat_value(
                chat,
                ["answer"]
            ),
            category,
            feedback,
            format_report_date(
                date_value
            ),
            chat_id
        ]

        for (
            column_index,
            value
        ) in enumerate(
            row_values,
            start=1
        ):

            cell = chat_sheet.cell(
                row=excel_row,
                column=column_index,
                value=value
            )

            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True
            )

    column_widths = {
        1: 8,
        2: 36,
        3: 55,
        4: 22,
        5: 20,
        6: 20,
        7: 38
    }

    for (
        column_index,
        width
    ) in column_widths.items():

        column_letter = get_column_letter(
            column_index
        )

        chat_sheet.column_dimensions[
            column_letter
        ].width = width

    chat_sheet.freeze_panes = (
        "A5"
    )

    statistics_sheet = (
        workbook.create_sheet(
            "İstatistikler"
        )
    )

    statistics_sheet.merge_cells(
        "A1:B1"
    )

    statistics_sheet["A1"] = (
        "Smart Support AI - Sistem İstatistikleri"
    )

    statistics_sheet["A1"].font = (
        title_font
    )

    statistics_sheet["A1"].alignment = Alignment(
        horizontal="center"
    )

    statistics_sheet["A3"] = (
        "İstatistik"
    )

    statistics_sheet["B3"] = (
        "Değer"
    )

    for cell in statistics_sheet[3]:

        cell.fill = header_fill
        cell.font = header_font

        cell.alignment = Alignment(
            horizontal="center"
        )

    statistic_rows = [
        (
            "Toplam soru",
            statistics[
                "total_questions"
            ]
        ),
        (
            "Olumlu geri bildirim",
            statistics[
                "positive_feedback"
            ]
        ),
        (
            "Olumsuz geri bildirim",
            statistics[
                "negative_feedback"
            ]
        ),
        (
            "Değerlendirilmeyen",
            statistics[
                "unanswered_feedback"
            ]
        ),
        (
            "En çok kullanılan kategori",
            statistics[
                "most_used_category_name"
            ]
        ),
        (
            "Kategori soru sayısı",
            statistics[
                "most_used_category_count"
            ]
        )
    ]

    for (
        row_index,
        statistic_row
    ) in enumerate(
        statistic_rows,
        start=4
    ):

        statistics_sheet.cell(
            row=row_index,
            column=1,
            value=statistic_row[0]
        )

        statistics_sheet.cell(
            row=row_index,
            column=2,
            value=statistic_row[1]
        )

    category_start_row = (
        len(statistic_rows)
        + 6
    )

    statistics_sheet.cell(
        row=category_start_row,
        column=1,
        value="Kategori"
    )

    statistics_sheet.cell(
        row=category_start_row,
        column=2,
        value="Soru Sayısı"
    )

    for column_index in range(
        1,
        3
    ):

        cell = statistics_sheet.cell(
            row=category_start_row,
            column=column_index
        )

        cell.fill = header_fill
        cell.font = header_font

        cell.alignment = Alignment(
            horizontal="center"
        )

    for (
        row_offset,
        category_data
    ) in enumerate(
        statistics[
            "category_statistics"
        ],
        start=1
    ):

        statistics_sheet.cell(
            row=(
                category_start_row
                + row_offset
            ),
            column=1,
            value=category_data[
                "name"
            ]
        )

        statistics_sheet.cell(
            row=(
                category_start_row
                + row_offset
            ),
            column=2,
            value=category_data[
                "count"
            ]
        )

    statistics_sheet.column_dimensions[
        "A"
    ].width = 38

    statistics_sheet.column_dimensions[
        "B"
    ].width = 22

    report_stream = BytesIO()

    workbook.save(
        report_stream
    )

    report_stream.seek(0)

    filename = (
        "smart_support_raporu_"
        + datetime.now().strftime(
            "%Y%m%d_%H%M%S"
        )
        + ".xlsx"
    )

    return send_file(
        report_stream,
        as_attachment=True,
        download_name=filename,
        mimetype=(
            "application/vnd.openxmlformats-"
            "officedocument.spreadsheetml.sheet"
        )
    )


# ==================================================
# PDF RAPORU
# ==================================================

@app.route("/admin/reports/pdf")
@admin_login_required
def download_pdf_report():
    """
    Sohbet geçmişini PDF raporu
    olarak indirir.
    """

    chat_history = prepare_chat_history(
        get_all_chats()
    )

    statistics = calculate_statistics(
        chat_history
    )

    (
        regular_font,
        bold_font
    ) = register_pdf_fonts()

    report_stream = BytesIO()

    document = SimpleDocTemplate(
        report_stream,
        pagesize=landscape(A4),
        rightMargin=1.2 * cm,
        leftMargin=1.2 * cm,
        topMargin=1.2 * cm,
        bottomMargin=1.2 * cm
    )

    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "SmartSupportTitle",
        parent=styles["Title"],
        fontName=bold_font,
        fontSize=18,
        leading=22,
        alignment=TA_CENTER,
        textColor=colors.HexColor(
            "#312E81"
        ),
        spaceAfter=12
    )

    normal_style = ParagraphStyle(
        "SmartSupportNormal",
        parent=styles["BodyText"],
        fontName=regular_font,
        fontSize=7.5,
        leading=10
    )

    bold_style = ParagraphStyle(
        "SmartSupportBold",
        parent=normal_style,
        fontName=bold_font
    )

    content = []

    content.append(
        Paragraph(
            "Smart Support AI - Teknik Destek Raporu",
            title_style
        )
    )

    content.append(
        Paragraph(
            "Rapor tarihi: {}".format(
                datetime.now().strftime(
                    "%d.%m.%Y %H:%M"
                )
            ),
            normal_style
        )
    )

    content.append(
        Spacer(
            1,
            0.4 * cm
        )
    )

    summary_data = [
        [
            Paragraph(
                "Toplam Soru",
                bold_style
            ),
            Paragraph(
                "Olumlu",
                bold_style
            ),
            Paragraph(
                "Olumsuz",
                bold_style
            ),
            Paragraph(
                "Değerlendirilmedi",
                bold_style
            ),
            Paragraph(
                "Popüler Kategori",
                bold_style
            )
        ],
        [
            str(
                statistics[
                    "total_questions"
                ]
            ),
            str(
                statistics[
                    "positive_feedback"
                ]
            ),
            str(
                statistics[
                    "negative_feedback"
                ]
            ),
            str(
                statistics[
                    "unanswered_feedback"
                ]
            ),
            Paragraph(
                str(
                    statistics[
                        "most_used_category_name"
                    ]
                ),
                normal_style
            )
        ]
    ]

    summary_table = Table(
        summary_data,
        colWidths=[
            3.2 * cm,
            3.2 * cm,
            3.2 * cm,
            3.8 * cm,
            7 * cm
        ]
    )

    summary_table.setStyle(
        TableStyle(
            [
                (
                    "BACKGROUND",
                    (0, 0),
                    (-1, 0),
                    colors.HexColor(
                        "#4F46E5"
                    )
                ),
                (
                    "TEXTCOLOR",
                    (0, 0),
                    (-1, 0),
                    colors.white
                ),
                (
                    "BACKGROUND",
                    (0, 1),
                    (-1, -1),
                    colors.HexColor(
                        "#EEF2FF"
                    )
                ),
                (
                    "ALIGN",
                    (0, 0),
                    (-1, -1),
                    "CENTER"
                ),
                (
                    "VALIGN",
                    (0, 0),
                    (-1, -1),
                    "MIDDLE"
                ),
                (
                    "GRID",
                    (0, 0),
                    (-1, -1),
                    0.5,
                    colors.HexColor(
                        "#CBD5E1"
                    )
                ),
                (
                    "TOPPADDING",
                    (0, 0),
                    (-1, -1),
                    8
                ),
                (
                    "BOTTOMPADDING",
                    (0, 0),
                    (-1, -1),
                    8
                )
            ]
        )
    )

    content.append(
        summary_table
    )

    content.append(
        Spacer(
            1,
            0.6 * cm
        )
    )

    table_data = [
        [
            Paragraph(
                "No",
                bold_style
            ),
            Paragraph(
                "Soru",
                bold_style
            ),
            Paragraph(
                "Cevap",
                bold_style
            ),
            Paragraph(
                "Kategori",
                bold_style
            ),
            Paragraph(
                "Geri Bildirim",
                bold_style
            ),
            Paragraph(
                "Tarih",
                bold_style
            )
        ]
    ]

    for (
        row_index,
        chat
    ) in enumerate(
        chat_history,
        start=1
    ):

        feedback = format_feedback(
            get_chat_value(
                chat,
                ["feedback"],
                ""
            )
        )

        category = get_chat_value(
            chat,
            [
                "category_name",
                "category"
            ],
            "Diğer"
        )

        date_value = get_chat_value(
            chat,
            [
                "created_at",
                "date",
                "timestamp"
            ],
            "-"
        )

        table_data.append(
            [
                str(row_index),
                Paragraph(
                    str(
                        get_chat_value(
                            chat,
                            ["question"]
                        )
                    ),
                    normal_style
                ),
                Paragraph(
                    str(
                        get_chat_value(
                            chat,
                            ["answer"]
                        )
                    ),
                    normal_style
                ),
                Paragraph(
                    str(category),
                    normal_style
                ),
                Paragraph(
                    feedback,
                    normal_style
                ),
                Paragraph(
                    format_report_date(
                        date_value
                    ),
                    normal_style
                )
            ]
        )

    if not chat_history:

        table_data.append(
            [
                "",
                Paragraph(
                    "Henüz sohbet kaydı bulunmuyor.",
                    normal_style
                ),
                "",
                "",
                "",
                ""
            ]
        )

    chat_table = Table(
        table_data,
        repeatRows=1,
        colWidths=[
            1 * cm,
            5 * cm,
            8.6 * cm,
            3.3 * cm,
            3.3 * cm,
            3.6 * cm
        ]
    )

    chat_table.setStyle(
        TableStyle(
            [
                (
                    "BACKGROUND",
                    (0, 0),
                    (-1, 0),
                    colors.HexColor(
                        "#4F46E5"
                    )
                ),
                (
                    "TEXTCOLOR",
                    (0, 0),
                    (-1, 0),
                    colors.white
                ),
                (
                    "ALIGN",
                    (0, 0),
                    (0, -1),
                    "CENTER"
                ),
                (
                    "VALIGN",
                    (0, 0),
                    (-1, -1),
                    "TOP"
                ),
                (
                    "GRID",
                    (0, 0),
                    (-1, -1),
                    0.35,
                    colors.HexColor(
                        "#CBD5E1"
                    )
                ),
                (
                    "ROWBACKGROUNDS",
                    (0, 1),
                    (-1, -1),
                    [
                        colors.white,
                        colors.HexColor(
                            "#F8FAFC"
                        )
                    ]
                ),
                (
                    "TOPPADDING",
                    (0, 0),
                    (-1, -1),
                    5
                ),
                (
                    "BOTTOMPADDING",
                    (0, 0),
                    (-1, -1),
                    5
                )
            ]
        )
    )

    content.append(
        chat_table
    )

    document.build(
        content
    )

    report_stream.seek(0)

    filename = (
        "smart_support_raporu_"
        + datetime.now().strftime(
            "%Y%m%d_%H%M%S"
        )
        + ".pdf"
    )

    return send_file(
        report_stream,
        as_attachment=True,
        download_name=filename,
        mimetype="application/pdf"
    )


# ==================================================
# VERİTABANI YEDEĞİ
# ==================================================

@app.route("/admin/reports/database")
@admin_login_required
def download_database_backup():
    """
    SQLite veritabanını indirir.
    """

    if not DATABASE_FILE.exists():

        flash(
            "Yedeklenecek veritabanı bulunamadı.",
            "error"
        )

        return redirect(
            url_for("admin_reports")
        )

    filename = (
        "smart_support_yedek_"
        + datetime.now().strftime(
            "%Y%m%d_%H%M%S"
        )
        + ".db"
    )

    return send_file(
        str(DATABASE_FILE),
        as_attachment=True,
        download_name=filename,
        mimetype="application/octet-stream"
    )


# ==================================================
# AYARLAR SAYFASI
# ==================================================

@app.route("/admin/settings")
@admin_login_required
def admin_settings():
    """
    Yönetici ayarları ve sistem
    bilgileri sayfasını gösterir.
    """

    settings = load_settings()

    system_versions = (
        get_system_versions()
    )

    support_items = load_questions()

    chat_history = prepare_chat_history(
        get_all_chats()
    )

    return render_template(
        "admin/settings.html",
        active_page="settings",
        admin_username=settings.get(
            "admin_username",
            "admin"
        ),
        system_versions=system_versions,
        total_knowledge_items=len(
            support_items
        ),
        total_chat_records=len(
            chat_history
        )
    )


# ==================================================
# YÖNETİCİ ŞİFRESİ DEĞİŞTİRME
# ==================================================

@app.route(
    "/admin/settings/password",
    methods=["POST"]
)
@admin_login_required
def change_admin_password():
    """
    Yönetici şifresini değiştirir.
    """

    current_password = request.form.get(
        "current_password",
        ""
    )

    new_password = request.form.get(
        "new_password",
        ""
    )

    confirm_password = request.form.get(
        "confirm_password",
        ""
    )

    settings = load_settings()

    if (
        current_password
        != settings["admin_password"]
    ):

        flash(
            "Mevcut yönetici şifresi yanlış.",
            "error"
        )

        return redirect(
            url_for("admin_settings")
        )

    if len(new_password) < 6:

        flash(
            "Yeni şifre en az 6 karakter olmalıdır.",
            "error"
        )

        return redirect(
            url_for("admin_settings")
        )

    if (
        new_password
        != confirm_password
    ):

        flash(
            "Yeni şifreler birbiriyle eşleşmiyor.",
            "error"
        )

        return redirect(
            url_for("admin_settings")
        )

    if (
        new_password
        == current_password
    ):

        flash(
            "Yeni şifre mevcut şifreyle aynı olamaz.",
            "error"
        )

        return redirect(
            url_for("admin_settings")
        )

    settings[
        "admin_password"
    ] = new_password

    if not save_settings(
        settings
    ):

        flash(
            "Şifre kaydedilirken bir hata oluştu.",
            "error"
        )

        return redirect(
            url_for("admin_settings")
        )

    flash(
        "Yönetici şifresi başarıyla değiştirildi.",
        "success"
    )

    return redirect(
        url_for("admin_settings")
    )


# ==================================================
# ADMIN SOHBET GEÇMİŞİNİ TEMİZLEME
# ==================================================

@app.route(
    "/admin/settings/clear-history",
    methods=["POST"]
)
@admin_login_required
def admin_clear_chat_history():
    """
    Sohbet geçmişini admin panelinden siler.
    """

    confirmation = request.form.get(
        "confirmation",
        ""
    )

    if confirmation != "confirm":

        flash(
            "Sohbet geçmişi silme işlemi onaylanmadı.",
            "error"
        )

        return redirect(
            url_for("admin_settings")
        )

    clear_all_chats()

    flash(
        "Tüm sohbet geçmişi başarıyla temizlendi.",
        "success"
    )

    return redirect(
        url_for("admin_settings")
    )


# ==================================================
# BİLGİ TABANINI JSON OLARAK İNDİRME
# ==================================================

@app.route(
    "/admin/settings/export-knowledge"
)
@admin_login_required
def export_knowledge_base():
    """
    Bilgi tabanını JSON yedeği
    olarak indirir.
    """

    support_items = load_questions()

    json_content = json.dumps(
        support_items,
        ensure_ascii=False,
        indent=4
    )

    report_stream = BytesIO(
        json_content.encode(
            "utf-8"
        )
    )

    filename = (
        "smart_support_bilgi_tabani_"
        + datetime.now().strftime(
            "%Y%m%d_%H%M%S"
        )
        + ".json"
    )

    return send_file(
        report_stream,
        as_attachment=True,
        download_name=filename,
        mimetype="application/json"
    )


# ==================================================
# BİLGİ TABANINI JSON DOSYASINDAN YÜKLEME
# ==================================================

@app.route(
    "/admin/settings/import-knowledge",
    methods=["POST"]
)
@admin_login_required
def import_knowledge_base():
    """
    JSON dosyasını bilgi tabanı
    olarak sisteme yükler.
    """

    uploaded_file = request.files.get(
        "knowledge_file"
    )

    if (
        uploaded_file is None
        or not uploaded_file.filename
    ):

        flash(
            "Lütfen bir JSON dosyası seçin.",
            "error"
        )

        return redirect(
            url_for("admin_settings")
        )

    if not uploaded_file.filename.lower().endswith(
        ".json"
    ):

        flash(
            "Yalnızca JSON dosyası yükleyebilirsiniz.",
            "error"
        )

        return redirect(
            url_for("admin_settings")
        )

    try:

        file_content = (
            uploaded_file.read().decode(
                "utf-8-sig"
            )
        )

        imported_data = json.loads(
            file_content
        )

    except UnicodeDecodeError:

        flash(
            "Dosyanın karakter kodlaması okunamadı.",
            "error"
        )

        return redirect(
            url_for("admin_settings")
        )

    except json.JSONDecodeError:

        flash(
            "Seçilen dosya geçerli bir JSON dosyası değil.",
            "error"
        )

        return redirect(
            url_for("admin_settings")
        )

    (
        is_valid,
        validation_message,
        validated_items
    ) = validate_imported_questions(
        imported_data
    )

    if not is_valid:

        flash(
            validation_message,
            "error"
        )

        return redirect(
            url_for("admin_settings")
        )

    if not save_questions(
        validated_items
    ):

        flash(
            "Bilgi tabanı kaydedilirken hata oluştu.",
            "error"
        )

        return redirect(
            url_for("admin_settings")
        )

    flash(
        "{} bilgi tabanı kaydı başarıyla içe aktarıldı.".format(
            len(
                validated_items
            )
        ),
        "success"
    )

    return redirect(
        url_for("admin_settings")
    )



# ==================================================
# KULLANICI GERİ BİLDİRİMİ
# ==================================================

@app.route(
    "/feedback/<chat_id>",
    methods=["POST"]
)
def save_feedback(chat_id):
    """
    Kullanıcının olumlu veya olumsuz
    geri bildirimini kaydeder.
    """

    feedback = request.form.get(
        "feedback",
        ""
    ).strip()

    if feedback not in [
        "positive",
        "negative"
    ]:

        flash(
            "Geçersiz geri bildirim.",
            "error"
        )

        return redirect(
            url_for("home")
        )

    update_feedback(
        chat_id,
        feedback
    )

    return redirect(
        url_for("home")
        + "#chat-"
        + chat_id
    )


# ==================================================
# YILDIZLI DEĞERLENDİRME
# ==================================================

@app.route(
    "/rating/<chat_id>",
    methods=["POST"]
)
def save_rating(chat_id):
    """
    Kullanıcının verdiği 1-5 arasındaki
    yıldız puanını kaydeder.
    """

    try:

        rating = int(
            request.form.get(
                "rating",
                "0"
            )
        )

    except (
        TypeError,
        ValueError
    ):

        flash(
            "Geçersiz puan.",
            "error"
        )

        return redirect(
            url_for("home")
        )

    if rating < 1 or rating > 5:

        flash(
            "Puan 1 ile 5 arasında olmalıdır.",
            "error"
        )

        return redirect(
            url_for("home")
        )

    rating_saved = update_rating(
        chat_id,
        rating
    )

    if not rating_saved:

        flash(
            "Değerlendirilecek sohbet bulunamadı.",
            "error"
        )

        return redirect(
            url_for("home")
        )

    flash(
        "Yıldız değerlendirmeniz kaydedildi.",
        "success"
    )

    return redirect(
        url_for("home")
        + "#chat-"
        + chat_id
    )


# ==================================================
# KULLANICI SOHBET GEÇMİŞİNİ TEMİZLEME
# ==================================================

@app.route(
    "/clear-history",
    methods=["POST"]
)
def clear_history():
    """
    Ana sayfadaki sohbet geçmişini temizler.
    """

    clear_all_chats()

    flash(
        "Sohbet geçmişi temizlendi.",
        "success"
    )

    return redirect(
        url_for("home")
    )


# ==================================================
# UYGULAMAYI ÇALIŞTIRMA
# ==================================================

if __name__ == "__main__":

    print("=" * 55)
    print("Smart Support AI başlatılıyor")
    print("=" * 55)

    print(
        "Şablon klasörü:",
        TEMPLATE_FOLDER
    )

    print(
        "Statik dosya klasörü:",
        STATIC_FOLDER
    )

    print(
        "Bilgi tabanı:",
        QUESTIONS_FILE
    )

    print(
        "Veritabanı:",
        DATABASE_FILE
    )

    print(
        "Ayarlar dosyası:",
        SETTINGS_FILE
    )

    print(
        "index.html:",
        (
            TEMPLATE_FOLDER
            / "index.html"
        ).exists()
    )

    print(
        "admin_login.html:",
        (
            TEMPLATE_FOLDER
            / "admin_login.html"
        ).exists()
    )

    print(
        "dashboard.html:",
        (
            TEMPLATE_FOLDER
            / "admin"
            / "dashboard.html"
        ).exists()
    )

    print(
        "knowledge.html:",
        (
            TEMPLATE_FOLDER
            / "admin"
            / "knowledge.html"
        ).exists()
    )

    print(
        "reports.html:",
        (
            TEMPLATE_FOLDER
            / "admin"
            / "reports.html"
        ).exists()
    )

    print(
        "settings.html:",
        (
            TEMPLATE_FOLDER
            / "admin"
            / "settings.html"
        ).exists()
    )

    print("=" * 55)

    app.run(
        debug=True
    )